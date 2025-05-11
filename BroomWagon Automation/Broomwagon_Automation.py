import os
import time
import logging
import shutil
import datetime
import re
import pandas as pd
import ctypes
import json
from openpyxl import load_workbook
from openpyxl.styles import numbers, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from win10toast import ToastNotifier
import subprocess
import sys

# === Auto-install missing dependencies ===
required_packages = [
    'pandas',
    'openpyxl',
    'win10toast'
]

for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        print(f"Package '{package}' not found. Installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# === Config ===
if getattr(sys, 'frozen', False):
    # Running as .exe (PyInstaller)
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    # Running as .py script
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # Directory where script is located

WATCH_FOLDER = SCRIPT_DIR  # Input files here
OUTPUT_FOLDER = os.path.join(SCRIPT_DIR, "Output")  # Processed files go here
CONFIG_FOLDER = os.path.join(SCRIPT_DIR, "config")  # Config folder for processed files, and driver config
LOG_DIR = os.path.join(SCRIPT_DIR, "logs")  # Logs folder
DEFAULT_DRIVERS = ["Hyndavi", "Sweta", "Shubham", "Joyita", "Shashwat", "Aman"]
CONFIG_PATH = os.path.join(CONFIG_FOLDER, "driver_status.json")
PROCESSED_FILES_LOG_PATH = os.path.join(CONFIG_FOLDER, "processed_files.log")


DRIVER_COLORS = {
    "Hyndavi": "e6b8b7",
    "Sweta": "b7dee8",
    "Shubham": "fabf8f",
    "Joyita": "c4d79b",
    "Shashwat": "c4bd97",
    "Aman": "b7dee8"
}

HEADER_COLORS = {
    "Driver": Font(color="FF0000"),
    "Ticket action": PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid"),
    "Internal action": PatternFill(start_color="92d050", end_color="92d050", fill_type="solid"),
    "reminder number\n(no number => #1)": PatternFill(start_color="fabf8f", end_color="fabf8f", fill_type="solid")
}

COLUMNS_EXPECTED = [
    'Ticket#', 'Age', 'Title', 'Created', 'Last Changed', 'Queue', 'State',
    'Priority', 'Customer ID', 'Service', 'SLA', 'Responsible',
    'UpdateTimeDestinationDate', 'SolutionTimeDestinationDate', 'Ticket-Ersteller'
]

# === Ensure required directories exist ===
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(CONFIG_FOLDER, exist_ok=True)

# Setup logging
log_filename = os.path.join(LOG_DIR, f"broomwagon_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")
logging.basicConfig(filename=log_filename, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
notifier = ToastNotifier()

# === Helpers ===
def get_current_week():
    today = datetime.date.today()
    return f"{today.isocalendar()[0]}-W{today.isocalendar()[1]:02}"

def get_drivers_for_week():
    week = get_current_week()
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "r") as f:
                config = json.load(f)
            drivers = config.get(week)
            if isinstance(drivers, list) and drivers:
                return drivers
        except Exception as e:
            logging.warning(f"Could not load config file: {e}")
    logging.info(f"No valid config found for {week}. Using default drivers.")
    return DEFAULT_DRIVERS[:]

def get_rotation_index():
    today = datetime.date.today()
    return (today.isocalendar()[1] - 1) % len(DEFAULT_DRIVERS)

# === Check if the file has been processed ===
def is_file_processed(filepath):
    if not os.path.exists(PROCESSED_FILES_LOG_PATH):
        return False
    with open(PROCESSED_FILES_LOG_PATH, "r") as f:
        processed_files = f.readlines()
    return any(filepath in line for line in processed_files)

# === Log the processed file ===
def log_processed_file(filepath):
    with open(PROCESSED_FILES_LOG_PATH, "a") as f:
        f.write(f"{filepath} - {datetime.datetime.now()}\n")
    logging.info(f"Logged processed file: {filepath}")

# === Processing function ===
def process_broomwagon_file(filepath):
    import time as t
    start = t.time()
    try:
        print(f"File: {os.path.basename(filepath)}")
        logging.info(f"Started processing file: {filepath}")

        # Check if the file has already been processed
        if is_file_processed(filepath):
            print(f"⚠️ This file has already been processed: {filepath}")
            logging.info(f"File has already been processed: {filepath}")
            return

        df = pd.read_excel(filepath, dtype={'Ticket#': str})

        if not all(col.strip() in df.columns for col in COLUMNS_EXPECTED):
            print("\033[91m❌ Error: Missing expected columns.\033[0m")
            logging.error("Missing expected columns. Aborting.")
            return

        base, ext = os.path.splitext(filepath)
        before_update_path = f"{base}_beforeupdate{ext}"
        shutil.move(filepath, before_update_path)
        filepath = os.path.join(OUTPUT_FOLDER, os.path.basename(f"{base}{ext}"))

        df.insert(0, 'Driver', '')
        ticket_index = df.columns.get_loc('Ticket#')
        df.insert(ticket_index + 1, 'Ticket action', '')
        df.insert(ticket_index + 2, 'Internal action', '')
        df.insert(ticket_index + 3, 'reminder number\n(no number => #1)', '')

        total_tickets = len(df)
        print(f"Total tickets: {total_tickets}\n")
        logging.info(f"Total tickets: {total_tickets}")

        DRIVERS = get_drivers_for_week()
        logging.info(f"Drivers used this week: {DRIVERS}")
        base_share = total_tickets // len(DRIVERS)
        extras = total_tickets % len(DRIVERS)
        start_idx = 0
        rotation = get_rotation_index()
        driver_summary = []

        for i in range(len(DRIVERS)):
            driver = DRIVERS[i]
            share = base_share + (1 if (i - rotation) % len(DRIVERS) < extras else 0)
            end_idx = start_idx + share
            df.loc[start_idx:end_idx - 1, 'Driver'] = driver
            driver_summary.append(f"• {driver}: {share} tickets")
            logging.info(f"Assigned {share} tickets to {driver} (Rows {start_idx + 2}-{end_idx + 1})")
            start_idx = end_idx

        df.to_excel(filepath, index=False)
        wb = load_workbook(filepath)
        ws = wb.active

        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'Ticket#':
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{row}"].number_format = numbers.FORMAT_TEXT
                break

        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'Driver':
                cell.font = HEADER_COLORS['Driver']
            elif cell.value in HEADER_COLORS:
                cell.fill = HEADER_COLORS[cell.value]
            cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

        for row in range(2, ws.max_row + 1):
            driver = ws[f"A{row}"].value
            color = DRIVER_COLORS.get(driver)
            if color:
                ws[f"A{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.freeze_panes = ws["B2"]

        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        special_columns = {
            'Ticket action': 35,
            'Internal action': 35,
            'reminder number\n(no number => #1)': 25
        }
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value in special_columns:
                ws.column_dimensions[get_column_letter(col_idx)].width = special_columns[cell.value]

        ws.row_dimensions[1].height = 35
        wb.save(filepath)

        print("\n✅ File processed successfully.\n")
        print("Ticket distribution:")
        for entry in driver_summary:
            print(entry)

        logging.info("Successfully completed processing.")
        print("\nProcessed file can be found in 'Output' folder")
        os.startfile(filepath)
        logging.info(f"Opened file: {filepath}")

        os.remove(before_update_path)
        logging.info(f"Deleted backup file: {before_update_path}")

        # Log the processed file
        log_processed_file(filepath)

    except Exception as e:
        print(f"\033[91m❌ Error occurred: {str(e)}\033[0m")
        logging.exception(f"Error occurred while processing file: {e}")
    finally:
        elapsed = round(t.time() - start, 2)
        print(f"\nTime taken: {elapsed} seconds")
        input("\nPress any key to continue...")

# === Folder monitoring ===
def monitor_folder():
    logging.info("Started BroomWagon watcher service.")
    logging.info(f"Watching folder: {WATCH_FOLDER}")

    if not os.path.exists(WATCH_FOLDER):
        print("\033[91m❌ Error: Watch folder does not exist.\033[0m")
        logging.error("Watch folder does not exist.")
        return

    # path for the driver script
    driver_script = os.path.join(SCRIPT_DIR, "Broomwagon Driver.py")

    while True:
        choice = input("Would you like to update the weekly driver availability? (y/n): ").strip().lower()
        if choice == 'y':
            if os.path.exists(driver_script):
                try:
                    #(f"Running driver script: {driver_script}")
                    subprocess.run([sys.executable, driver_script], check=True)
                    logging.info(f"Driver script executed: {driver_script}")
                    break  
                except Exception as e:
                    print(f"\033[91mWarning: Could not launch driver script: {e}\033[0m")
                    logging.warning(f"Could not launch driver script: {e}")
                    break
            else:
                print("\033[91mWarning: Driver script not found. Continuing without update.\033[0m")
                logging.warning("Driver script not found. Continuing without update.")
                break
        elif choice == 'n':
            break  # Exit the loop without running the script
        else:
            print("\033[91mInvalid input. Please enter 'y' or 'n'.\033[0m")

    files = [f for f in os.listdir(WATCH_FOLDER) if f.endswith(".xlsx")]
    if not files:
        print("\033[91m❌ Error: No files found to process.\033[0m")
        logging.error("No files found to process.")
        return

    for file in files:
        process_broomwagon_file(os.path.join(WATCH_FOLDER, file))

if __name__ == "__main__":
    monitor_folder()
